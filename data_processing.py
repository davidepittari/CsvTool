import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.styles import Font, Alignment
from openpyxl.styles import PatternFill

def reset_cell_styles(sheet, start_row, num_rows, start_col, num_cols):
    """Resets the font, size, and alignment of cells in a specific range."""
    for row in range(start_row, start_row + num_rows):
        for col in range(start_col, start_col + num_cols):
            cell = sheet.cell(row=row, column=col)
            cell.font = Font()  # Reset font to default
            cell.alignment = Alignment()  # Reset alignment to default


def check_sheet_exists(book, sheet_name):
    """Check if a sheet exists in the Excel file."""
    if sheet_name not in book.sheetnames:
        raise ValueError(f"The sheet '{sheet_name}' is not present in the Excel file.")

def insert_data_into_sheet(sheet, data, start_row, dest_column):
    """Inserts data into the specified sheet starting from the given row and column."""
    for i, value in enumerate(data):
        row = start_row + i
        sheet.cell(row=row, column=dest_column, value=value)
        
def set_row_height(sheet, start_row, num_rows, height):
    """Set the height of rows in a specific range."""
    for row in range(start_row, start_row + num_rows):
        sheet.row_dimensions[row].height = height
        
def unmerge_cells_in_rows(sheet, start_row, num_rows):
    """Unmerges cells in specific rows."""
    end_row = start_row + num_rows
    merged_ranges = list(sheet.merged_cells.ranges)
    for merged_range in merged_ranges:
        min_row, max_row = merged_range.bounds[1], merged_range.bounds[3]
        if min_row >= start_row and max_row < end_row:
            sheet.unmerge_cells(str(merged_range))


def format_columns(sheet, start_row, dest_column, data_length):
    """Formats the column background to green and adds borders."""
    green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    border_style = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )
    
    for i in range(start_row, start_row + data_length):
        cell = sheet.cell(row=i, column=dest_column)
        cell.fill = green_fill
        cell.border = border_style

def format_row(sheet, row_number):
    """Formats the entire row background to yellow and adds borders."""
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    border_style = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )
    
    for cell in sheet[row_number]:
        cell.fill = yellow_fill
        cell.border = border_style

def apply_row_formats(sheet, start_row, num_rows):
    """Apply specific formatting to cells in the specified rows."""
    # Define fills for different ranges
    gray_fill = PatternFill(start_color="A6A6A6", end_color="A6A6A6", fill_type="solid")  # Gray
    green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")  # Green
    pink_fill = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")  # Pink

    # Define border style
    border_style = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )

    for row in range(start_row, start_row + num_rows):
        # Apply gray fill and borders to column A
        cell = sheet.cell(row=row, column=1)
        cell.fill = gray_fill
        cell.border = border_style

        # Apply green fill to columns B to K
        for col in range(2, 12):  # Columns B to K
            sheet.cell(row=row, column=col).fill = green_fill
            sheet.cell(row=row, column=col).border = border_style

        # Apply pink fill to columns L and M
        for col in range(12, 14):  # Columns L and M
            sheet.cell(row=row, column=col).fill = pink_fill
            sheet.cell(row=row, column=col).border = border_style


def process_data(csv_file, excel_file, excel_file_final, column_vars, column_dest_vars, start_row_var, sheet_name_var, dropdown_option, dropdown_cell_var, is_final):
    df = pd.read_csv(csv_file)
    book = load_workbook(excel_file)
    final_book = load_workbook(excel_file_final, data_only=True)  # Load final file

    check_sheet_exists(book, sheet_name_var.get())

    sheet = book[sheet_name_var.get()]

    start_row = start_row_var.get()
    num_rows = len(df)

    # Unmerge cells only in the specific rows where data will be inserted
    unmerge_cells_in_rows(sheet, start_row, num_rows)

    # Set the row height, reset cell styles, and apply row-specific formats
    set_row_height(sheet, start_row, num_rows, 37.5)
    reset_cell_styles(sheet, start_row, num_rows, 1, sheet.max_column)
    apply_row_formats(sheet, start_row, num_rows)

    dropdown_cell = dropdown_cell_var.get()
    sheet[dropdown_cell] = dropdown_option

    dropdown_row = sheet[dropdown_cell].row

    for column, var in column_vars.items():
        if var.get():
            dest_column_letter = column_dest_vars[column].get().strip().upper()
            dest_col_idx = ord(dest_column_letter) - ord('A') + 1
            insert_data_into_sheet(sheet, df[column].tolist(), start_row, dest_col_idx)

            # Apply formatting to copied columns
            format_columns(sheet, start_row, dest_col_idx, len(df[column]))

    # Apply formatting to row containing dropdown option
    format_row(sheet, dropdown_row)
    
    if is_final:
     # Copy content from the final file
        for sheet_name in final_book.sheetnames:
            final_sheet = final_book[sheet_name]
            # Copy all cells from final_sheet to the end of the target sheet
            max_row_target = sheet.max_row
            max_col_final = final_sheet.max_column
            max_row_final = final_sheet.max_row
            num_rows_final = max_row_final
            
            for row in final_sheet.iter_rows(min_row=1, max_row=max_row_final, max_col=max_col_final):
                for cell in row:
                    new_cell = sheet.cell(row=max_row_target + cell.row, column=cell.column, value=cell.value)
                    if cell.has_style:
                        new_cell.font = cell.font.copy()
                        new_cell.fill = cell.fill.copy()
                        new_cell.border = cell.border.copy()
                        new_cell.alignment = cell.alignment.copy()

        set_row_height(sheet, max_row_target, num_rows_final, 37.5)

    book.save(excel_file)



